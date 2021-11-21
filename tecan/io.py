from datetime import datetime
from typing import Iterable, Union, BinaryIO, List, Optional

import openpyxl
import numpy as np
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from tecan.data import InstrumentMetadata, SingleAssay, PlateMetadata, AssayMetadata, Plate, TimeSeries

from pint import Quantity, UndefinedUnitError


class TecanReader:
    TS_INFINITE_FORMAT = '%m/%d/%Y %I:%M:%S %p'
    TS_SPARK_FORMAT = '%Y-%m-%d %H:%M:%S'

    @classmethod
    def read(cls, file: Union[str, BinaryIO]) -> Plate:
        """
        Reads a Tecan Infinite plate from an .xlsx file.
        :param file: string or a binary file-like object
        :return: TecanInfinitePlate
        """
        ws = cls._get_ws(file)
        plate_metadata, instrument_metadata, assay_metadata = cls._read_header(ws)
        assays = cls._get_assays(file, ws, assay_metadata)
        return Plate(assays, plate_metadata, instrument_metadata)

    @classmethod
    def _get_ws(cls, file: Union[str, BinaryIO]) -> Worksheet:
        wb = openpyxl.load_workbook(file, read_only=True)
        return wb.active

    @classmethod
    def _get_assays(cls, file: Union[str, BinaryIO], ws: Worksheet,
                    assay_metadata: Iterable[AssayMetadata]) -> Iterable[SingleAssay]:

        a_list = cls._find_assays(ws)
        assay_data = [cls._read_assay(file, ws, cell[0], cell[1])
                      for cell in zip(a_list, a_list[1:] + [None])]
        return [SingleAssay(data, metadata) if len(data.columns) == 1 else TimeSeries(data, metadata)
                for data, metadata in zip(assay_data, assay_metadata)]

    @classmethod
    def _clean_datetime(cls, dt: Union[str, datetime], formats: List[str]) -> Optional[datetime]:
        """
        Attempts to create a datetime object from a variety of formats.
        :param dt: string or existing datetime object
        :param formats: list of datetime formats to try
        :return:
        """

        result = None
        if isinstance(dt, str):
            for f in formats:
                try:
                    result = datetime.strptime(dt, f)
                    break
                except ValueError:
                    pass
        elif isinstance(dt, datetime):
            result = dt

        return result

    @classmethod
    def _clean_time(cls, t: str) -> Optional[datetime]:
        """
        Converts time string into a datetime object.
        :param t: string
        :return: datetime
        """
        formats = ['%I:%M:%S %p', '%H:%M:%S', '%H:%M']
        return cls._clean_datetime(t, formats)

    @classmethod
    def _clean_date(cls, d):
        formats = ['%m/%d/%Y', '%Y/%m/%d', '%m-%d-%Y', '%Y-%m-%d']
        return cls._clean_datetime(d, formats)


    @classmethod
    def _read_header(cls, ws):

        date = cls._clean_date(cls._find_exact(ws, 'Date:', col_shift=range(1, 5), min_col=1, max_col=1))
        time = cls._clean_time(cls._find_exact(ws, 'Time:', col_shift=range(1, 5), min_col=1, max_col=1))

        instrument_metadata = {
            'application': cls._find_beginning(ws, 'Application: ', min_col=1, max_col=1),
            'name': cls._find_beginning(ws, 'Device: ', min_col=1, max_col=1),
            'firmware': cls._find_beginning(ws, 'Firmware: ', min_col=1, max_col=1),
            'serial_number': cls._find_beginning(ws, 'Serial number: ', min_col=5, max_col=5),
            'system': cls._find_exact(ws, 'System', col_shift=4, min_col=1, max_col=1),
            'user': cls._find_exact(ws, 'User', col_shift=4, min_col=1, max_col=1),
            'session_start': datetime.strptime(date.strftime('%Y-%m-%d ') + time.strftime('%H:%M:%S'),
                                               cls.TS_SPARK_FORMAT)
        }

        def _to_time(t):
            return cls._clean_datetime(t, [cls.TS_INFINITE_FORMAT, cls.TS_SPARK_FORMAT])

        plate_metadata = {
            'type': cls._find_exact(ws, 'Plate', col_shift=4, min_col=1, max_col=1),
            'start': _to_time(cls._find_header(ws, 'Start Time', col_shift=range(1, 5), min_col=1, max_col=1)),
            'end': _to_time(cls._find_header(ws, 'End Time', col_shift=range(1, 5), min_col=1, max_col=1)),
        }

        assay_metadata = [cls._read_label(ws, label) for label in cls._find_labels(ws)]

        return PlateMetadata(plate_metadata), InstrumentMetadata(instrument_metadata), assay_metadata

    @classmethod
    def _get_label_mode(cls, ws, label):
        mode = None
        for i in range(1, 5):
            mode = ws.cell(row=label.row, column=label.column + i).value
            if mode:
                break
        return mode

    @classmethod
    def _get_label_name(cls, ws, label):
        mode = None
        for i in range(1, 5):
            mode = ws.cell(row=label.row, column=label.column + i).value
            if mode:
                break
        return mode

    @classmethod
    def _find_labels(cls, ws):
        labels = []
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if str(cell.value).lower() == "mode" and str(cls._get_label_mode(ws, cell)).lower() != 'kinetic':
                    labels.append(cell)
        return labels

    @classmethod
    def _read_label(cls, ws, label):

        metadata = {
            'mode': cls._get_label_mode(ws, label)
        }
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=label.row + 1, max_row=label.row + 21):
            for cell in row:
                if (str(cell.value).lower() == "mode" or
                        str(cell.value).lower() == "part of plate" or
                        not cell.value):
                    break
                key = str(cell.value).lower().replace(' ', '_').replace('-', '_').replace('(', '').replace(')', '')
                value = ws.cell(row=cell.row, column=cell.column + 4).value
                unit = ws.cell(row=cell.row, column=cell.column + 5).value
                if unit is None:
                    metadata[key] = value
                else:
                    try:
                        metadata[key] = Quantity(value, unit)
                    except (UndefinedUnitError, ValueError):
                        metadata[key] = (value, unit)
            else:
                continue
            break
        return AssayMetadata(metadata)

    @classmethod
    def _find_assays(cls, ws):
        assays = []
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == "Cycle Nr." or cell.value == "<>":
                    assays.append(cell)
        return assays

    @classmethod
    def _read_assay(cls, file, ws, cell, last=None):
        nrows = last.row - cell.row if last is not None else None
        if cell.value == "Cycle Nr.":
            name = ws.cell(row=cell.row-1, column=cell.column).value
            raw = pd.read_excel(file, skiprows=cell.row-1, index_col=0, nrows=nrows) \
                    .replace('OVER', np.inf) \
                    .apply(pd.to_numeric, errors='coerce')
            times = pd.Index(pd.to_timedelta(raw.iloc[0], unit='S')).rename('time')
            temps = pd.Index(raw.iloc[1]).rename('temperature').astype('float')
            data = raw.rename_axis(index='well', columns=raw.index.name) \
                      .drop(raw[~raw.index.to_series().astype('str').str.match("[A-Z]{1,2}[0-9]{1,3}")].index)
            index = data.index.to_series().str.extract('([A-Z]+)([0-9]+)') \
                        .rename({0: 'row', 1: 'column'}, axis='columns').apply(pd.to_numeric, errors='ignore') \
                        .set_index(['row', 'column']).index
            columns = pd.MultiIndex.from_arrays(
                [data.columns.rename('cycle').astype('int'),
                 pd.Series([name for _ in data.columns], name='label'), times, temps])
            data = pd.DataFrame(data.values, index=index, columns=columns) \
                     .drop(columns[np.isnat(columns.get_level_values(2))], axis='columns')
        elif cell.value == "<>":
            temperature = float(ws.cell(row=cell.row-1, column=cell.column+1).value
                                  .replace("Temperature: ", "").replace(" °C", ""))
            raw = pd.read_excel(file, skiprows=cell.row-1, header=0, index_col=0, nrows=nrows) \
                    .replace('OVER', np.inf).apply(pd.to_numeric, errors='coerce')
            data = raw.loc[
                    [i for i in raw.index if str(i).isupper() and len(i) == 1],
                    [c for c in raw.columns if not str(c).startswith('Unnamed: ')]] \
                .rename_axis(index='row', columns='column') \
                .melt(var_name='column', ignore_index=False).set_index('column', append=True).dropna()
            columns = pd.MultiIndex.from_arrays([
                pd.Series([1], name='cycle'),
                pd.Series(['Assay'], name='label'),
                pd.Series([np.timedelta64(0, 's')], name='time'),
                pd.Series([temperature], name='temperature')
            ])
            data = pd.DataFrame(data.values, index=data.index, columns=columns)
        else:
            raise ValueError

        return data

    @classmethod
    def _get_shifted_value(cls, ws, c, col_shift):
        if isinstance(col_shift, Iterable):
            for s in col_shift:
                v = ws.cell(row=c.row, column=c.column + s).value
                if v:
                    return v
        else:
            return ws.cell(row=c.row, column=c.column + col_shift).value

        return None

    @classmethod
    def _find_exact(cls, ws, value, col_shift: Union[int, Iterable, None] = 1, **opts):
        return cls._find_cell(ws, lambda c: c.value == value,
                              lambda c: cls._get_shifted_value(ws, c, col_shift), **opts)

    @classmethod
    def _find_header(cls, ws, value, col_shift: Union[int, Iterable, None] = 1, **opts):
        return cls._find_cell(ws, lambda c: str(c.value).strip(': \t\n\r') == value,
                              lambda c: cls._get_shifted_value(ws, c, col_shift), **opts)

    @classmethod
    def _find_beginning(cls, ws, value, **opts):
        return cls._find_cell(ws, lambda c: str(c.value).startswith(value),
                              lambda c: c.value.replace(value, ''), **opts)

    @classmethod
    def _find_cell(cls, ws, condition, transform=None, **opts):
        for row in ws.iter_rows(**opts):
            for cell in row:
                if condition(cell):
                    if transform is None:
                        return cell.value
                    else:
                        return transform(cell)
        return None
